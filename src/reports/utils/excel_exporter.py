"""Excel export functionality with frozen panes and formatting"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self, log_func):
        self.log = log_func
    
    def export_to_excel(self, data_dict, output_path):
        """
        Export multiple DataFrames to Excel with separate tabs
        
        Args:
            data_dict: Dictionary where keys are tab names, values are DataFrames
            output_path: Path for the output Excel file
        """
        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            for tab_name, df in data_dict.items():
                self.log(f"Processing tab: {tab_name}")
                self.log(f"  Type: {type(df)}")
                if hasattr(df, 'shape'):
                    self.log(f"  Shape: {df.shape}")
                else:
                    self.log(f"  No shape attribute (not a DataFrame?)")
                if df is None or len(df) == 0:
                    self.log(f"  Skipping tab '{tab_name}' (empty or None)")
                    continue
                
                print(f"DEBUG: Processing tab: {tab_name}")
                
                # Fix timezone issues before processing
                df_clean = self._fix_timezone_issues(df)
                
                # Create worksheet
                ws = wb.create_sheet(title=self._clean_sheet_name(tab_name))
                
                # Add data
                for r in dataframe_to_rows(df_clean, index=False, header=True):
                    ws.append(r)
                
                # Format the sheet
                self._format_worksheet(ws, df_clean, tab_name)
                
                print(f"DEBUG: Added sheet: {tab_name} ({len(df_clean)} rows)")
            
            # Save the workbook
            wb.save(output_path)
            print(f"DEBUG: Excel export complete: {len(data_dict)} tabs created")
            return output_path
            
        except Exception as e:
            self.log(f"Error creating Excel file: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            return None
    
    def _fix_timezone_issues(self, df):
        """Fix timezone issues that cause Excel export problems"""
        df_copy = df.copy()
        
        for col in df_copy.columns:
            # Check if column contains datetime data
            if df_copy[col].dtype.name.startswith('datetime'):
                try:
                    # If it's timezone-aware, convert to timezone-naive
                    if hasattr(df_copy[col].dtype, 'tz') and df_copy[col].dtype.tz is not None:
                        df_copy[col] = df_copy[col].dt.tz_localize(None)
                        print(f"DEBUG: Removed timezone from column: {col}")
                except Exception as e:
                    print(f"DEBUG: Warning: Could not fix timezone for column {col}: {str(e)}")
            
            # Handle object columns that might contain datetime objects
            elif df_copy[col].dtype == 'object':
                try:
                    # Check if it contains datetime-like objects
                    sample = df_copy[col].dropna().head(3)
                    if len(sample) > 0:
                        first_val = sample.iloc[0]
                        if hasattr(first_val, 'tz') and first_val.tz is not None:
                            # Convert timezone-aware datetime objects to timezone-naive
                            df_copy[col] = pd.to_datetime(df_copy[col]).dt.tz_localize(None)
                            print(f"DEBUG: Fixed timezone in object column: {col}")
                except:
                    # If conversion fails, leave the column as-is
                    pass
        
        return df_copy
    
    def _clean_sheet_name(self, name):
        """Clean sheet name to be Excel-compatible"""
        # Excel sheet names can't exceed 31 chars and can't contain certain characters
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        clean_name = name
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        return clean_name[:31]
    
    def _format_worksheet(self, ws, df, tab_name=None):
        """Apply formatting to worksheet"""
        self.log(f"Formatting worksheet for tab: {tab_name}")
        self.log(f"  Worksheet max_row: {ws.max_row}, max_column: {ws.max_column}")
        if ws.max_row == 0:
            self.log(f"  Worksheet is empty, skipping formatting.")
            return
        
        print(f"DEBUG: Formatting worksheet for tab: {tab_name}")
        print(f"DEBUG: DataFrame has _gender_timeline_formatting: {hasattr(df, '_gender_timeline_formatting')}")
        print(f"DEBUG: DataFrame has _red_score_data: {hasattr(df, '_red_score_data')}")
        
        # Format header row
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        header_font = Font(bold=True)
        
        for cell in ws[1]:  # First row
            cell.fill = header_fill
            cell.font = header_font

        # Apply gender timeline conditional formatting if this is a gender timeline sheet
        if hasattr(df, '_gender_timeline_formatting') and hasattr(df, '_red_score_data'):
            print("DEBUG: About to apply gender timeline formatting")
            self._apply_gender_timeline_formatting(ws, df)
        else:
            print("DEBUG: Skipping gender timeline formatting - flags not found")
        
        # Set frozen panes safely
        from openpyxl.utils import get_column_letter
        num_cols = len(df.columns)
        
        # For gender timeline, freeze more columns to keep FLW info visible
        if hasattr(df, '_gender_timeline_formatting'):
            freeze_at_col = 4  # Freeze flw_id, flw_name, opportunity_name
        elif 'flw_id' in df.columns or any('id' in col.lower() for col in df.columns[:3]):
            freeze_at_col = 4 if num_cols > 3 else 3
        else:
            freeze_at_col = 3
        
        # Only freeze if the column exists
        if num_cols >= freeze_at_col:
            freeze_col_letter = get_column_letter(freeze_at_col)
            ws.freeze_panes = f'{freeze_col_letter}2'
        else:
            ws.freeze_panes = 'A2'
        
        # Auto-adjust column widths (with reasonable limits)
        if ws.max_column > 0:
            for column in ws.columns:
                max_length = 0
                # Defensive: skip if column is empty
                if not column or not hasattr(column[0], 'column'):
                    self.log(f"  Skipping empty or invalid column in tab {tab_name}")
                    continue
                if hasattr(column[0], 'column'):
                    column_index = column[0].column
                elif hasattr(column[0], 'col_idx'):
                    column_index = column[0].col_idx
                else:
                    # fallback or skip
                    continue
                column_letter = column_index
                #column_letter = get_column_letter(column_index)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        self.log(f"  Error measuring cell value length in column {column_letter}: {e}")
                # Set width with reasonable bounds
                adjusted_width = min(max(max_length + 2, 10), 50)
                try:
<<<<<<< HEAD
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with reasonable bounds
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_gender_timeline_formatting(self, ws, df):
        """Apply conditional formatting to gender timeline sheets"""
        print("DEBUG: Starting _apply_gender_timeline_formatting")
        
        try:
            red_score_df = df._red_score_data
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            
            print(f"DEBUG: Red score dataframe shape: {red_score_df.shape}")
            
            # Find week columns (columns that start with 'Week_')
            week_columns = [col for col in df.columns if col.startswith('Week_')]
            print(f"DEBUG: Found {len(week_columns)} week columns")
            
            # Get column indices for week columns
            col_indices = {}
            for idx, col_name in enumerate(df.columns):
                if col_name in week_columns:
                    col_indices[col_name] = idx + 1  # +1 because Excel is 1-indexed
            
            print(f"DEBUG: Applying red score formatting to {len(week_columns)} week columns for {len(df)} FLWs")
            
            # Debug: Check how many red scores we have
            total_red_scores = 0
            for col_name in week_columns:
                if col_name in red_score_df.columns:
                    red_count = (red_score_df[col_name] == True).sum()
                    total_red_scores += red_count
                    if red_count > 0:
                        print(f"DEBUG: Week {col_name}: {red_count} red scores")
            
            print(f"DEBUG: Total red scores across all weeks: {total_red_scores}")
            
            if total_red_scores == 0:
                print("DEBUG: No red scores found - no cells will be highlighted")
                return
            
            # Apply formatting to cells where red_score_data is True
            red_cells_count = 0
            for row_idx in range(len(df)):
                for col_name in week_columns:
                    if (col_name in red_score_df.columns and 
                        row_idx < len(red_score_df) and 
                        red_score_df.iloc[row_idx][col_name] == True):
                        
                        excel_row = row_idx + 2  # +2 for header row (1-indexed)
                        excel_col = col_indices[col_name]
                        
                        cell = ws.cell(row=excel_row, column=excel_col)
                        cell.fill = red_fill
                        red_cells_count += 1
                        
                        # Debug: Log first few red cells
                        if red_cells_count <= 5:
                            print(f"DEBUG: Applied red formatting to row {excel_row}, col {excel_col} ({col_name})")
            
            print(f"DEBUG: Applied red formatting to {red_cells_count} cells with red scores")
            
        except Exception as e:
            print(f"DEBUG: Warning: Could not apply gender timeline formatting: {str(e)}")
            import traceback
            traceback.print_exc()
            # Continue without formatting rather than failing
=======
                    ws.column_dimensions[column_letter].width = adjusted_width
                except Exception as e:
                    self.log(f"  Error setting column width for {column_letter} in tab {tab_name}: {e}")
>>>>>>> 289f4321edd46570cefbc03b59cb3b145a1e81d5
