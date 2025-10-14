"""
ML Fraud Detection Report

Trains machine learning models on aggregated FLW features to detect fraudulent data patterns.
Generates unified ranking scores and comprehensive analysis across multiple models.
"""

import tkinter as tk
from tkinter import ttk
import pandas as pd
import numpy as np
import os
from pathlib import Path
from datetime import datetime
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import classification_report, roc_auc_score, precision_recall_curve, roc_curve
import warnings
warnings.filterwarnings('ignore')

from .base_report import BaseReport


class MLFraudDetectionReport(BaseReport):
    """
    ML-based fraud detection using aggregated FLW features.
    Provides unified ranking scores and threshold analysis.
    """
    
    @staticmethod
    def setup_parameters(parent_frame):
        """Set up GUI parameters for ML fraud detection"""
        
        # Train/test split ratio
        ttk.Label(parent_frame, text="Test split ratio:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        parent_frame.test_split_var = tk.StringVar(value="0.2")
        ttk.Entry(parent_frame, textvariable=parent_frame.test_split_var, width=10).grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Models to include
        ttk.Label(parent_frame, text="Models to train:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        parent_frame.include_rf_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(parent_frame, text="Random Forest", variable=parent_frame.include_rf_var).grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        
        parent_frame.include_lr_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(parent_frame, text="Logistic Regression", variable=parent_frame.include_lr_var).grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Class balancing
        ttk.Label(parent_frame, text="Handle class imbalance:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        parent_frame.balance_method_var = tk.StringVar(value="class_weight")
        balance_combo = ttk.Combobox(parent_frame, textvariable=parent_frame.balance_method_var,
                                   values=["none", "class_weight"], state="readonly", width=15)
        balance_combo.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Random state for reproducibility
        ttk.Label(parent_frame, text="Random seed:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
        parent_frame.random_state_var = tk.StringVar(value="42")
        ttk.Entry(parent_frame, textvariable=parent_frame.random_state_var, width=10).grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)

    def generate(self):
        """Generate ML fraud detection analysis"""
        
        # Get parameters
        test_split = float(self.get_parameter_value('test_split', 0.2))
        include_rf = self.get_parameter_value('include_rf', True)
        include_lr = self.get_parameter_value('include_lr', True)
        balance_method = self.get_parameter_value('balance_method', 'class_weight')
        random_state = int(self.get_parameter_value('random_state', 42))
        use_date_subdir = self.get_parameter_value('use_date_subdir', True)
        
        self.log("Starting ML fraud detection analysis...")
        
        # Load and prepare data
        X, y, feature_names, flw_metadata = self._prepare_ml_data()
        self.log(f"Prepared {len(X)} FLWs with {len(feature_names)} features")
        self.log(f"Class distribution - Real: {(y == 0).sum()}, Fake: {(y == 1).sum()}")
        
        # Split data
        X_train, X_test, y_train, y_test, meta_train, meta_test = train_test_split(
            X, y, flw_metadata, test_size=test_split, stratify=y, random_state=random_state
        )
        
        self.log(f"Split - Train: {len(X_train)}, Test: {len(X_test)}")
        
        # Train models
        models = {}
        predictions = {}
        
        if include_rf:
            self.log("Training Random Forest...")
            rf_model = self._train_random_forest(X_train, y_train, balance_method, random_state)
            models['random_forest'] = rf_model
            predictions['rf_prob'] = rf_model.predict_proba(X)[:, 1]
            
        if include_lr:
            self.log("Training Logistic Regression...")
            lr_model, scaler = self._train_logistic_regression(X_train, y_train, balance_method, random_state)
            models['logistic_regression'] = (lr_model, scaler)
            X_scaled = scaler.transform(X)
            predictions['lr_prob'] = lr_model.predict_proba(X_scaled)[:, 1]
        
        # Calculate composite score
        if len(predictions) > 1:
            predictions['composite_prob'] = np.mean(list(predictions.values()), axis=0)
        else:
            predictions['composite_prob'] = list(predictions.values())[0]
        
        # Create output directory
        if use_date_subdir:
            today = datetime.now().strftime("%Y_%m_%d")
            output_dir = os.path.join(self.output_dir, f"ml_fraud_detection_{today}")
        else:
            output_dir = self.output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate all outputs
        output_files = []
        
        # 1. All FLWs Ranked
        ranked_df = self._create_ranked_flws(flw_metadata, predictions, models, X, feature_names)
        
        # 2. Model Performance
        performance_df = self._evaluate_models(models, X_train, X_test, y_train, y_test, feature_names)
        
        # 3. Feature Importance
        importance_df = self._extract_feature_importance(models, feature_names)
        
        # 4. Threshold Analysis
        threshold_df = self._analyze_thresholds(y, predictions['composite_prob'])
        
        # Save as Excel file with multiple tabs
        total_flws = len(flw_metadata)
        real_flws = (flw_metadata['classification'] == 'real').sum()
        fake_flws = (flw_metadata['classification'] == 'fake').sum()
        
        excel_file = os.path.join(output_dir, f"ml_fraud_detection_{total_flws}flws_{real_flws}real_{fake_flws}fake.xlsx")
        
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            ranked_df.to_excel(writer, sheet_name="All_FLWs_Ranked", index=False)
            performance_df.to_excel(writer, sheet_name="Model_Performance", index=False)
            importance_df.to_excel(writer, sheet_name="Feature_Importance", index=False)
            threshold_df.to_excel(writer, sheet_name="Threshold_Analysis", index=False)
            
            # Add conditional formatting to highlight fake FLWs
            self._format_ranked_sheet(writer, ranked_df)
        
        output_files.append(excel_file)
        
        # Summary statistics
        self._log_summary_stats(ranked_df, threshold_df)
        
        self.log(f"ML fraud detection complete! Generated: {os.path.basename(excel_file)}")
        return output_files
    
    def _prepare_ml_data(self):
        """Prepare data for ML training"""
        
        # Load the aggregated features
        df = self.df.copy()
        
        # Validate required columns
        required_cols = ['flw_id', 'classification']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        # Separate metadata from features
        metadata_cols = ['flw_id', 'flw_name', 'opportunity_id', 'opportunity_name', 'visits', 'classification']
        flw_metadata = df[metadata_cols].copy()
        
        # Auto-detect feature columns (everything else)
        feature_cols = [col for col in df.columns if col not in metadata_cols]
        
        self.log(f"Auto-detected {len(feature_cols)} feature columns")
        self.log(f"Features: {', '.join(feature_cols[:10])}{'...' if len(feature_cols) > 10 else ''}")
        
        # Extract features and target
        X = df[feature_cols].copy()
        
        # Handle missing values (replace -1 sentinel values with NaN, then impute)
        X = X.replace(-1, np.nan)
        
        # Simple imputation: fill numeric with median, categorical with mode
        for col in X.columns:
            if X[col].dtype in ['object', 'category']:
                X[col] = X[col].fillna(X[col].mode().iloc[0] if not X[col].mode().empty else 'unknown')
            else:
                X[col] = X[col].fillna(X[col].median())
        
        # Create target variable (0 = real, 1 = fake)
        y = (df['classification'] == 'fake').astype(int)
        
        return X, y, feature_cols, flw_metadata
    
    def _train_random_forest(self, X_train, y_train, balance_method, random_state):
        """Train Random Forest model"""
        
        rf_params = {
            'n_estimators': 100,
            'max_depth': 10,
            'min_samples_split': 5,
            'min_samples_leaf': 2,
            'random_state': random_state
        }
        
        if balance_method == 'class_weight':
            rf_params['class_weight'] = 'balanced'
        
        rf_model = RandomForestClassifier(**rf_params)
        rf_model.fit(X_train, y_train)
        
        return rf_model
    
    def _train_logistic_regression(self, X_train, y_train, balance_method, random_state):
        """Train Logistic Regression model"""
        
        # Scale features for logistic regression
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        
        lr_params = {
            'random_state': random_state,
            'max_iter': 1000
        }
        
        if balance_method == 'class_weight':
            lr_params['class_weight'] = 'balanced'
        
        lr_model = LogisticRegression(**lr_params)
        lr_model.fit(X_train_scaled, y_train)
        
        return lr_model, scaler
    
    def _create_ranked_flws(self, flw_metadata, predictions, models, X, feature_names):
        """Create ranked list of all FLWs with fraud scores"""
        
        ranked_df = flw_metadata.copy()
        
        # Add prediction probabilities
        for pred_name, probs in predictions.items():
            ranked_df[pred_name] = probs
        
        # Add rank based on composite score
        ranked_df['fraud_rank'] = ranked_df['composite_prob'].rank(method='dense', ascending=False).astype(int)
        
        # Add top contributing features for each FLW
        if 'random_forest' in models:
            top_features = self._get_top_contributing_features(models['random_forest'], X, feature_names, n_features=3)
            ranked_df['top_contributing_features'] = top_features
        
        # Sort by fraud rank
        ranked_df = ranked_df.sort_values('fraud_rank').reset_index(drop=True)
        
        # Reorder columns for readability
        base_cols = ['fraud_rank', 'flw_id', 'classification', 'composite_prob']
        pred_cols = [col for col in ranked_df.columns if col.endswith('_prob') and col != 'composite_prob']
        meta_cols = ['flw_name', 'opportunity_id', 'opportunity_name', 'visits']
        feature_cols = ['top_contributing_features'] if 'top_contributing_features' in ranked_df.columns else []
        
        final_cols = base_cols + pred_cols + meta_cols + feature_cols
        final_cols = [col for col in final_cols if col in ranked_df.columns]
        
        return ranked_df[final_cols]
    
    def _get_top_contributing_features(self, rf_model, X, feature_names, n_features=3):
        """Get top contributing features for each FLW using Random Forest"""
        
        # Get feature importances from the model
        importances = rf_model.feature_importances_
        
        # For each FLW, find features with highest values weighted by importance
        top_features_list = []
        
        for idx, row in X.iterrows():
            # Calculate contribution score for each feature (value * importance)
            contributions = row.values * importances
            
            # Get top N features
            top_indices = np.argsort(contributions)[-n_features:][::-1]
            top_feature_names = [feature_names[i] for i in top_indices]
            top_contributions = [contributions[i] for i in top_indices]
            
            # Format as string
            feature_str = "; ".join([f"{name}={contrib:.3f}" for name, contrib in 
                                   zip(top_feature_names, top_contributions)])
            top_features_list.append(feature_str)
        
        return top_features_list
    
    def _evaluate_models(self, models, X_train, X_test, y_train, y_test, feature_names):
        """Evaluate model performance"""
        
        performance_data = []
        
        for model_name, model_info in models.items():
            if model_name == 'random_forest':
                model = model_info
                X_test_processed = X_test
            elif model_name == 'logistic_regression':
                model, scaler = model_info
                X_test_processed = scaler.transform(X_test)
            
            # Predictions
            y_pred_proba = model.predict_proba(X_test_processed)[:, 1]
            y_pred = (y_pred_proba > 0.5).astype(int)
            
            # Metrics
            try:
                auc_score = roc_auc_score(y_test, y_pred_proba)
            except:
                auc_score = np.nan
            
            performance_data.append({
                'model': model_name,
                'test_auc': auc_score,
                'test_samples': len(y_test),
                'test_positive_rate': y_test.mean(),
            })
        
        return pd.DataFrame(performance_data)
    
    def _extract_feature_importance(self, models, feature_names):
        """Extract and rank feature importance from models"""
        
        importance_data = []
        
        if 'random_forest' in models:
            rf_importance = models['random_forest'].feature_importances_
            
            # Create a dataframe to properly handle ranking with ties
            rf_df = pd.DataFrame({
                'feature': feature_names,
                'rf_importance': rf_importance
            })
            
            # Rank by importance (descending), handling ties with min method
            rf_df['rf_rank'] = rf_df['rf_importance'].rank(method='min', ascending=False).astype(int)
            
            for _, row in rf_df.iterrows():
                importance_data.append({
                    'feature': row['feature'],
                    'rf_importance': row['rf_importance'],
                    'rf_rank': row['rf_rank']
                })
        
        if 'logistic_regression' in models:
            lr_model, _ = models['logistic_regression']
            lr_coef = np.abs(lr_model.coef_[0])  # Absolute value of coefficients
            lr_importance_sorted = np.argsort(lr_coef)[::-1]
            
            for i, feature in enumerate(feature_names):
                existing_row = next((row for row in importance_data if row['feature'] == feature), None)
                
                coef_abs = lr_coef[i]
                rank = np.where(lr_importance_sorted == i)[0][0] + 1
                
                if existing_row:
                    existing_row['lr_importance'] = coef_abs
                    existing_row['lr_rank'] = rank
                else:
                    importance_data.append({
                        'feature': feature,
                        'lr_importance': coef_abs,
                        'lr_rank': rank
                    })
        
        importance_df = pd.DataFrame(importance_data)
        
        # Calculate composite rank
        rank_cols = [col for col in importance_df.columns if col.endswith('_rank')]
        if len(rank_cols) > 0:
            importance_df['avg_rank'] = importance_df[rank_cols].mean(axis=1)
            importance_df = importance_df.sort_values('avg_rank')
        
        return importance_df
    
    def _analyze_thresholds(self, y_true, y_scores):
        """Analyze different threshold values for classification"""
        
        # Calculate precision-recall curve
        precisions, recalls, thresholds = precision_recall_curve(y_true, y_scores)
        
        # Add threshold=1.0 case
        thresholds = np.append(thresholds, 1.0)
        
        threshold_data = []
        
        for i, threshold in enumerate(thresholds):
            predictions = (y_scores >= threshold).astype(int)
            
            # Calculate metrics
            tp = np.sum((predictions == 1) & (y_true == 1))
            fp = np.sum((predictions == 1) & (y_true == 0))
            tn = np.sum((predictions == 0) & (y_true == 0))
            fn = np.sum((predictions == 0) & (y_true == 1))
            
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            
            # False positive rate (flagged real / total real)
            fpr = fp / (fp + tn) if (fp + tn) > 0 else 0
            
            threshold_data.append({
                'threshold': threshold,
                'precision': precision,
                'recall': recall,
                'false_positive_rate': fpr,
                'flagged_real_pct': fpr * 100,
                'caught_fake_pct': recall * 100,
                'total_flagged': tp + fp,
                'total_real': tn + fp,
                'total_fake': tp + fn
            })
        
        threshold_df = pd.DataFrame(threshold_data)
        
        # Add specific analysis for user's questions
        analysis_rows = []
        
        # Question 1: "If flag at most 10% of real data, how much fake data caught?"
        max_10pct_real = threshold_df[threshold_df['flagged_real_pct'] <= 10]
        if not max_10pct_real.empty:
            best_10pct = max_10pct_real.loc[max_10pct_real['caught_fake_pct'].idxmax()]
            analysis_rows.append({
                'analysis': 'Flag =10% real data',
                'threshold': best_10pct['threshold'],
                'flagged_real_pct': best_10pct['flagged_real_pct'],
                'caught_fake_pct': best_10pct['caught_fake_pct']
            })
        
        # Question 2: "If catch at least 50% of fake data, how little real data flagged?"
        min_50pct_fake = threshold_df[threshold_df['caught_fake_pct'] >= 50]
        if not min_50pct_fake.empty:
            best_50pct = min_50pct_fake.loc[min_50pct_fake['flagged_real_pct'].idxmin()]
            analysis_rows.append({
                'analysis': 'Catch =50% fake data',
                'threshold': best_50pct['threshold'],
                'flagged_real_pct': best_50pct['flagged_real_pct'],
                'caught_fake_pct': best_50pct['caught_fake_pct']
            })
        
        # Add analysis rows at the top
        if analysis_rows:
            analysis_df = pd.DataFrame(analysis_rows)
            threshold_df = pd.concat([analysis_df, threshold_df], ignore_index=True)
        
        return threshold_df
    
    def _log_summary_stats(self, ranked_df, threshold_df):
        """Log summary statistics"""
        
        total_flws = len(ranked_df)
        real_flws = (ranked_df['classification'] == 'real').sum()
        fake_flws = (ranked_df['classification'] == 'fake').sum()
        
        self.log(f"Summary: {total_flws} FLWs analyzed ({real_flws} real, {fake_flws} fake)")
        
        # Most suspicious real FLW
        most_suspicious_real = ranked_df[ranked_df['classification'] == 'real'].iloc[0]
        self.log(f"Most suspicious real FLW: {most_suspicious_real['flw_id']} (rank {most_suspicious_real['fraud_rank']}, score {most_suspicious_real['composite_prob']:.3f})")
        
        # Threshold analysis highlights
        analysis_rows = threshold_df[threshold_df['analysis'].notna()]
        for _, row in analysis_rows.iterrows():
            self.log(f"{row['analysis']}: threshold={row['threshold']:.3f} ? {row['flagged_real_pct']:.1f}% real flagged, {row['caught_fake_pct']:.1f}% fake caught")
    
    def _format_ranked_sheet(self, writer, ranked_df):
        """Apply conditional formatting to highlight fake FLWs"""
        
        try:
            from openpyxl.styles import PatternFill
            
            workbook = writer.book
            worksheet = writer.sheets["All_FLWs_Ranked"]
            
            # Light yellow fill for fake FLWs
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            
            # Find the classification column
            classification_col = None
            for col_idx, col_name in enumerate(ranked_df.columns, 1):
                if col_name == 'classification':
                    classification_col = col_idx
                    break
            
            if classification_col is not None:
                # Apply formatting to rows where classification = 'fake'
                for row_idx, (_, row_data) in enumerate(ranked_df.iterrows(), 2):  # Start at row 2 (after header)
                    if row_data['classification'] == 'fake':
                        # Apply yellow fill to entire row
                        for col_idx in range(1, len(ranked_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill
                            
        except ImportError:
            # If openpyxl formatting isn't available, continue without formatting
            pass
        except Exception as e:
            # If any other formatting error occurs, log but don't fail
            self.log(f"Note: Could not apply Excel formatting - {str(e)}")

    @classmethod
    def create_for_automation(cls, df, output_dir, test_split=0.2, include_rf=True, 
                            include_lr=True, balance_method='class_weight', random_state=42,
                            use_date_subdir=True):
        """
        Create report instance for automated pipeline use (no GUI)
        
        Args:
            df: DataFrame with aggregated FLW features (from MLFeatureAggregationReport)
            output_dir: Directory for output files
            test_split: Fraction of data for testing (0-1)
            include_rf: Whether to train Random Forest model
            include_lr: Whether to train Logistic Regression model
            balance_method: Class imbalance handling - 'none' or 'class_weight'
            random_state: Random seed for reproducibility
            use_date_subdir: Whether to create dated subdirectory (default True)
        
        Returns:
            MLFraudDetectionReport instance ready to call generate()
        """
        # Create a minimal mock params frame that mimics tkinter variables
        class MockVar:
            def __init__(self, value):
                self._value = value
            def get(self):
                return self._value
            def set(self, value):
                self._value = value
        
        class MockParamsFrame:
            def __init__(self):
                self.test_split_var = MockVar(str(test_split))
                self.include_rf_var = MockVar(include_rf)
                self.include_lr_var = MockVar(include_lr)
                self.balance_method_var = MockVar(balance_method)
                self.random_state_var = MockVar(str(random_state))
                self.use_date_subdir_var = MockVar(use_date_subdir)
        
        mock_frame = MockParamsFrame()
        
        # Add dummy log callback to match BaseReport signature
        def dummy_log(msg):
            pass
        
        # Create instance - match BaseReport's expected parameters
        instance = cls(
            df=df,
            output_dir=output_dir,
            log_callback=dummy_log,
            params_frame=mock_frame
        )
        
        return instance

