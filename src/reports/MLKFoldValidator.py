"""
ML K-Fold Cross-Validation

Performs k-fold cross-validation for fraud detection models to assess overfitting
and generate robust rankings. Tracks predictions across all folds and uses the
median AUC fold for final rankings.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from sklearn.model_selection import StratifiedKFold
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import roc_auc_score
import warnings
warnings.filterwarnings('ignore')


class MLKFoldValidator:
    """
    K-fold cross-validation for ML fraud detection.
    Provides robust performance estimates and identifies consistently suspicious FLWs.
    
    Generates 5 Excel tabs:
    1. Ranked_FLWs_Median_Fold: All FLWs ranked by fraud score from median-performing fold
    2. Fold_Metrics: Performance metrics (AUC, etc.) for each fold
    3. Feature_Importance: Average feature importance across all folds
    4. Prediction_Variance: FLWs sorted by prediction inconsistency across folds
    5. Threshold_Analysis: Trade-offs between audit burden and fraud detection at different thresholds
    """
    
    def __init__(self, df, output_dir, n_folds=10, include_rf=True, include_lr=True,
                 balance_method='class_weight', random_state=42, log_callback=None,
                 features_for_ml='all'):
        """
        Initialize k-fold validator
        
        Args:
            df: DataFrame with aggregated FLW features
            output_dir: Directory for output files
            n_folds: Number of folds (default 10)
            include_rf: Whether to train Random Forest
            include_lr: Whether to train Logistic Regression
            balance_method: Class imbalance handling
            random_state: Random seed for reproducibility
            log_callback: Optional logging function
            features_for_ml: Either "all" or list of feature names to use for training
        """
        self.df = df
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.n_folds = n_folds
        self.include_rf = include_rf
        self.include_lr = include_lr
        self.balance_method = balance_method
        self.random_state = random_state
        self.log_callback = log_callback or print
        self.features_for_ml = features_for_ml
        
        # Will be populated during validation
        self.fold_results = []
        self.all_predictions = {}
        self.median_fold_idx = None
        
    def log(self, msg):
        """Log message"""
        if self.log_callback:
            self.log_callback(msg)
    
    def run_kfold_validation(self):
        """Run k-fold cross-validation and generate reports"""
        
        self.log(f"Starting {self.n_folds}-fold cross-validation...")
        
        # Prepare data
        X, y, feature_names, flw_metadata = self._prepare_data()
        self.log(f"Prepared {len(X)} FLWs with {len(feature_names)} features")
        self.log(f"Class distribution - Real: {(y == 0).sum()}, Fake: {(y == 1).sum()}")
        
        # Create stratified k-fold splits
        skf = StratifiedKFold(n_splits=self.n_folds, shuffle=True, random_state=self.random_state)
        
        # Store predictions for all FLWs across all folds
        if self.include_rf:
            self.all_predictions['rf_prob'] = np.zeros((len(X), self.n_folds))
        if self.include_lr:
            self.all_predictions['lr_prob'] = np.zeros((len(X), self.n_folds))
        
        # Run each fold
        for fold_idx, (train_idx, test_idx) in enumerate(skf.split(X, y), 1):
            self.log(f"\nFold {fold_idx}/{self.n_folds}")
            self.log("-" * 40)
            
            fold_result = self._train_and_evaluate_fold(
                X, y, feature_names, train_idx, test_idx, fold_idx
            )
            
            self.fold_results.append(fold_result)
            
            self.log(f"  Test AUC: {fold_result['test_auc']:.4f}")
        
        # Find median AUC fold
        self._find_median_fold()
        
        # Generate reports
        output_files = self._generate_reports(flw_metadata, feature_names, X, y)
        
        self.log(f"\n? K-fold validation complete!")
        return output_files
    
    def _prepare_data(self):
        """Prepare data for ML (same as MLFraudDetectionReport)"""
        
        df = self.df.copy()
        
        # Validate required columns
        required_cols = ['flw_id', 'classification']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        # Separate metadata from features
        metadata_cols = ['flw_id', 'flw_name', 'opportunity_id', 'opportunity_name', 'visits', 'classification']
        flw_metadata = df[metadata_cols].copy()
        
        # Auto-detect ALL available feature columns
        all_feature_cols = [col for col in df.columns if col not in metadata_cols]
        
        # Apply feature selection
        if self.features_for_ml == 'all':
            feature_cols = all_feature_cols
            self.log(f"Using all {len(feature_cols)} available features")
        elif isinstance(self.features_for_ml, list):
            # Validate that requested features exist
            missing_features = [f for f in self.features_for_ml if f not in all_feature_cols]
            if missing_features:
                raise ValueError(f"Requested features not found in data: {missing_features}")
            feature_cols = self.features_for_ml
            self.log(f"Using custom feature set: {len(feature_cols)} features")
            self.log(f"Features: {', '.join(feature_cols)}")
        else:
            raise ValueError(f"features_for_ml must be 'all' or a list of feature names, got: {self.features_for_ml}")
        
        # Extract features
        X = df[feature_cols].copy()
        
        # Handle missing values
        X = X.replace(-1, np.nan)
        for col in X.columns:
            if X[col].dtype in ['object', 'category']:
                X[col] = X[col].fillna(X[col].mode().iloc[0] if not X[col].mode().empty else 'unknown')
            else:
                X[col] = X[col].fillna(X[col].median())
        
        # Create target variable (0 = real, 1 = fake)
        y = (df['classification'] == 'fake').astype(int).values
        
        return X.values, y, feature_cols, flw_metadata
    
    def _train_and_evaluate_fold(self, X, y, feature_names, train_idx, test_idx, fold_idx):
        """Train models and evaluate on one fold"""
        
        X_train, X_test = X[train_idx], X[test_idx]
        y_train, y_test = y[train_idx], y[test_idx]
        
        fold_result = {
            'fold': fold_idx,
            'train_size': len(train_idx),
            'test_size': len(test_idx),
            'test_idx': test_idx,
            'feature_importance': {}
        }
        
        # Train Random Forest
        if self.include_rf:
            rf_model = self._train_random_forest(X_train, y_train)
            
            # Predict on ALL FLWs (not just test set)
            rf_prob_all = rf_model.predict_proba(X)[:, 1]
            self.all_predictions['rf_prob'][:, fold_idx - 1] = rf_prob_all
            
            # Evaluate on test set only
            rf_prob_test = rf_prob_all[test_idx]
            fold_result['rf_test_auc'] = roc_auc_score(y_test, rf_prob_test)
            fold_result['feature_importance']['rf'] = rf_model.feature_importances_
            
            self.log(f"  RF Test AUC: {fold_result['rf_test_auc']:.4f}")
        
        # Train Logistic Regression
        if self.include_lr:
            lr_model, scaler = self._train_logistic_regression(X_train, y_train)
            
            # Predict on ALL FLWs
            X_scaled_all = scaler.transform(X)
            lr_prob_all = lr_model.predict_proba(X_scaled_all)[:, 1]
            self.all_predictions['lr_prob'][:, fold_idx - 1] = lr_prob_all
            
            # Evaluate on test set only
            lr_prob_test = lr_prob_all[test_idx]
            fold_result['lr_test_auc'] = roc_auc_score(y_test, lr_prob_test)
            fold_result['feature_importance']['lr'] = np.abs(lr_model.coef_[0])
            
            self.log(f"  LR Test AUC: {fold_result['lr_test_auc']:.4f}")
        
        # Calculate composite AUC (average of models)
        test_aucs = []
        if self.include_rf:
            test_aucs.append(fold_result['rf_test_auc'])
        if self.include_lr:
            test_aucs.append(fold_result['lr_test_auc'])
        
        fold_result['test_auc'] = np.mean(test_aucs)
        
        return fold_result
    
    def _train_random_forest(self, X_train, y_train):
        """Train Random Forest model"""
        
        rf_params = {
            'n_estimators': 100,
            'max_depth': 10,
            'min_samples_split': 5,
            'min_samples_leaf': 2,
            'random_state': self.random_state
        }
        
        if self.balance_method == 'class_weight':
            rf_params['class_weight'] = 'balanced'
        
        rf_model = RandomForestClassifier(**rf_params)
        rf_model.fit(X_train, y_train)
        
        return rf_model
    
    def _train_logistic_regression(self, X_train, y_train):
        """Train Logistic Regression model"""
        
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        
        lr_params = {
            'random_state': self.random_state,
            'max_iter': 1000
        }
        
        if self.balance_method == 'class_weight':
            lr_params['class_weight'] = 'balanced'
        
        lr_model = LogisticRegression(**lr_params)
        lr_model.fit(X_train_scaled, y_train)
        
        return lr_model, scaler
    
    def _find_median_fold(self):
        """Find fold with median test AUC"""
        
        test_aucs = [fold['test_auc'] for fold in self.fold_results]
        median_auc = np.median(test_aucs)
        
        # Find fold closest to median
        auc_diffs = [abs(auc - median_auc) for auc in test_aucs]
        self.median_fold_idx = np.argmin(auc_diffs)
        
        self.log(f"\nMedian AUC: {median_auc:.4f}")
        self.log(f"Median fold: {self.median_fold_idx + 1} (AUC: {test_aucs[self.median_fold_idx]:.4f})")
    
    def _generate_reports(self, flw_metadata, feature_names, X, y):
        """Generate comprehensive reports with 5 Excel tabs"""
        
        self.log("\nGenerating reports...")
        
        # Create output directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_subdir = self.output_dir / f"kfold_validation_{timestamp}"
        output_subdir.mkdir(exist_ok=True)
        
        # Calculate composite predictions (average across models)
        composite_preds = np.mean(list(self.all_predictions.values()), axis=0)
        
        # 1. Ranked FLWs from median fold
        ranked_df = self._create_ranked_flws(flw_metadata, composite_preds, feature_names, X)
        
        # 2. Fold-by-fold metrics
        metrics_df = self._create_metrics_summary()
        
        # 3. Feature importance stability
        importance_df = self._create_feature_importance_summary(feature_names)
        
        # 4. Per-FLW prediction variance
        variance_df = self._create_prediction_variance(flw_metadata, composite_preds, y)
        
        # 5. Threshold analysis (using median fold predictions)
        median_composite = composite_preds[:, self.median_fold_idx]
        threshold_df = self._create_threshold_analysis(y, median_composite)
        
        # Log summary
        self._log_summary_stats(ranked_df, threshold_df)
        
        # 5. Threshold analysis (using median fold predictions)
        median_composite = composite_preds[:, self.median_fold_idx]
        threshold_df = self._create_threshold_analysis(y, median_composite)
        
        # Save to Excel
        total_flws = len(flw_metadata)
        real_flws = (flw_metadata['classification'] == 'real').sum()
        fake_flws = (flw_metadata['classification'] == 'fake').sum()
        
        excel_file = output_subdir / f"kfold_results_{total_flws}flws_{real_flws}real_{fake_flws}fake.xlsx"
        
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            ranked_df.to_excel(writer, sheet_name="Ranked_FLWs_Median_Fold", index=False)
            metrics_df.to_excel(writer, sheet_name="Fold_Metrics", index=False)
            importance_df.to_excel(writer, sheet_name="Feature_Importance", index=False)
            variance_df.to_excel(writer, sheet_name="Prediction_Variance", index=False)
            threshold_df.to_excel(writer, sheet_name="Threshold_Analysis", index=False)
            
            # Apply formatting
            self._format_excel(writer, ranked_df)
        
        self.log(f"Generated: {excel_file.name}")
        
        return [str(excel_file)]
    
    def _create_ranked_flws(self, flw_metadata, composite_preds, feature_names, X):
        """Create ranked FLW list from median fold predictions"""
        
        ranked_df = flw_metadata.copy()
        
        # Use median fold predictions
        median_composite = composite_preds[:, self.median_fold_idx]
        ranked_df['composite_prob'] = median_composite
        
        # Add individual model predictions from median fold if available
        if self.include_rf:
            ranked_df['rf_prob'] = self.all_predictions['rf_prob'][:, self.median_fold_idx]
        if self.include_lr:
            ranked_df['lr_prob'] = self.all_predictions['lr_prob'][:, self.median_fold_idx]
        
        # Calculate prediction variance across all folds
        ranked_df['prediction_std'] = np.std(composite_preds, axis=1)
        ranked_df['prediction_min'] = np.min(composite_preds, axis=1)
        ranked_df['prediction_max'] = np.max(composite_preds, axis=1)
        
        # Add fraud rank
        ranked_df['fraud_rank'] = ranked_df['composite_prob'].rank(method='dense', ascending=False).astype(int)
        
        # Sort by rank
        ranked_df = ranked_df.sort_values('fraud_rank').reset_index(drop=True)
        
        # Reorder columns
        base_cols = ['fraud_rank', 'flw_id', 'classification', 'composite_prob', 
                     'prediction_std', 'prediction_min', 'prediction_max']
        pred_cols = [col for col in ranked_df.columns if col.endswith('_prob') and col != 'composite_prob']
        meta_cols = ['flw_name', 'opportunity_id', 'opportunity_name', 'visits']
        
        final_cols = base_cols + pred_cols + meta_cols
        final_cols = [col for col in final_cols if col in ranked_df.columns]
        
        return ranked_df[final_cols]
    
    def _create_metrics_summary(self):
        """Create fold-by-fold metrics summary"""
        
        metrics_data = []
        
        for fold_result in self.fold_results:
            row = {
                'fold': fold_result['fold'],
                'test_size': fold_result['test_size'],
                'test_auc': fold_result['test_auc']
            }
            
            if self.include_rf:
                row['rf_test_auc'] = fold_result['rf_test_auc']
            if self.include_lr:
                row['lr_test_auc'] = fold_result['lr_test_auc']
            
            metrics_data.append(row)
        
        metrics_df = pd.DataFrame(metrics_data)
        
        # Add summary statistics
        summary_row = {
            'fold': 'MEAN',
            'test_size': metrics_df['test_size'].mean(),
            'test_auc': metrics_df['test_auc'].mean()
        }
        if self.include_rf:
            summary_row['rf_test_auc'] = metrics_df['rf_test_auc'].mean()
        if self.include_lr:
            summary_row['lr_test_auc'] = metrics_df['lr_test_auc'].mean()
        
        std_row = {
            'fold': 'STD',
            'test_size': 0,
            'test_auc': metrics_df['test_auc'].std()
        }
        if self.include_rf:
            std_row['rf_test_auc'] = metrics_df['rf_test_auc'].std()
        if self.include_lr:
            std_row['lr_test_auc'] = metrics_df['lr_test_auc'].std()
        
        median_row = {
            'fold': 'MEDIAN',
            'test_size': metrics_df['test_size'].median(),
            'test_auc': metrics_df['test_auc'].median()
        }
        if self.include_rf:
            median_row['rf_test_auc'] = metrics_df['rf_test_auc'].median()
        if self.include_lr:
            median_row['lr_test_auc'] = metrics_df['lr_test_auc'].median()
        
        summary_df = pd.DataFrame([summary_row, std_row, median_row])
        metrics_df = pd.concat([metrics_df, summary_df], ignore_index=True)
        
        return metrics_df
    
    def _create_feature_importance_summary(self, feature_names):
        """Create feature importance summary across all folds"""
        
        importance_data = []
        
        for feat_idx, feat_name in enumerate(feature_names):
            row = {'feature': feat_name}
            
            if self.include_rf:
                rf_importances = [fold['feature_importance']['rf'][feat_idx] 
                                 for fold in self.fold_results]
                row['rf_importance_mean'] = np.mean(rf_importances)
                row['rf_importance_std'] = np.std(rf_importances)
            
            if self.include_lr:
                lr_importances = [fold['feature_importance']['lr'][feat_idx] 
                                 for fold in self.fold_results]
                row['lr_importance_mean'] = np.mean(lr_importances)
                row['lr_importance_std'] = np.std(lr_importances)
            
            importance_data.append(row)
        
        importance_df = pd.DataFrame(importance_data)
        
        # Sort by average importance
        if self.include_rf:
            importance_df = importance_df.sort_values('rf_importance_mean', ascending=False)
        elif self.include_lr:
            importance_df = importance_df.sort_values('lr_importance_mean', ascending=False)
        
        importance_df = importance_df.reset_index(drop=True)
        
        return importance_df
    
    def _create_prediction_variance(self, flw_metadata, composite_preds, y):
        """Analyze prediction variance across folds for each FLW"""
        
        variance_df = flw_metadata.copy()
        
        # Calculate statistics across folds
        variance_df['pred_mean'] = np.mean(composite_preds, axis=1)
        variance_df['pred_std'] = np.std(composite_preds, axis=1)
        variance_df['pred_min'] = np.min(composite_preds, axis=1)
        variance_df['pred_max'] = np.max(composite_preds, axis=1)
        variance_df['pred_range'] = variance_df['pred_max'] - variance_df['pred_min']
        
        # Count how many folds flagged as suspicious (>0.5 threshold)
        variance_df['times_flagged'] = (composite_preds > 0.5).sum(axis=1)
        
        # Add actual label
        variance_df['is_fake'] = y
        
        # Sort by prediction variance (most inconsistent first)
        variance_df = variance_df.sort_values('pred_std', ascending=False).reset_index(drop=True)
        
        return variance_df
    
    def _create_threshold_analysis(self, y_true, y_scores):
        """Analyze different threshold values for classification (using median fold predictions)"""
        
        from sklearn.metrics import precision_recall_curve
        
        # Calculate precision-recall curve
        precisions, recalls, thresholds = precision_recall_curve(y_true, y_scores)
        
        # Add threshold=1.0 case
        thresholds = np.append(thresholds, 1.0)
        
        threshold_data = []
        
        for i, threshold in enumerate(thresholds):
            predictions = (y_scores >= threshold).astype(int)
            
            # Calculate metrics
            tp = np.sum((predictions == 1) & (y_true == 1))
            fp = np.sum((predictions == 1) & (y_true == 0))
            tn = np.sum((predictions == 0) & (y_true == 0))
            fn = np.sum((predictions == 0) & (y_true == 1))
            
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            
            # False positive rate (flagged real / total real)
            fpr = fp / (fp + tn) if (fp + tn) > 0 else 0
            
            threshold_data.append({
                'threshold': threshold,
                'precision': precision,
                'recall': recall,
                'false_positive_rate': fpr,
                'flagged_real_pct': fpr * 100,
                'caught_fake_pct': recall * 100,
                'total_flagged': tp + fp,
                'total_real': tn + fp,
                'total_fake': tp + fn
            })
        
        threshold_df = pd.DataFrame(threshold_data)
        
        # Add specific analysis for user's questions
        analysis_rows = []
        
        # Question 1: "If flag at most 10% of real data, how much fake data caught?"
        max_10pct_real = threshold_df[threshold_df['flagged_real_pct'] <= 10]
        if not max_10pct_real.empty:
            best_10pct = max_10pct_real.loc[max_10pct_real['caught_fake_pct'].idxmax()]
            analysis_rows.append({
                'analysis': 'Flag =10% real data',
                'threshold': best_10pct['threshold'],
                'flagged_real_pct': best_10pct['flagged_real_pct'],
                'caught_fake_pct': best_10pct['caught_fake_pct'],
                'precision': best_10pct['precision'],
                'recall': best_10pct['recall']
            })
        
        # Question 2: "If catch at least 50% of fake data, how little real data flagged?"
        min_50pct_fake = threshold_df[threshold_df['caught_fake_pct'] >= 50]
        if not min_50pct_fake.empty:
            best_50pct = min_50pct_fake.loc[min_50pct_fake['flagged_real_pct'].idxmin()]
            analysis_rows.append({
                'analysis': 'Catch =50% fake data',
                'threshold': best_50pct['threshold'],
                'flagged_real_pct': best_50pct['flagged_real_pct'],
                'caught_fake_pct': best_50pct['caught_fake_pct'],
                'precision': best_50pct['precision'],
                'recall': best_50pct['recall']
            })
        
        # Question 3: "If catch at least 90% of fake data, how little real data flagged?"
        min_90pct_fake = threshold_df[threshold_df['caught_fake_pct'] >= 90]
        if not min_90pct_fake.empty:
            best_90pct = min_90pct_fake.loc[min_90pct_fake['flagged_real_pct'].idxmin()]
            analysis_rows.append({
                'analysis': 'Catch =90% fake data',
                'threshold': best_90pct['threshold'],
                'flagged_real_pct': best_90pct['flagged_real_pct'],
                'caught_fake_pct': best_90pct['caught_fake_pct'],
                'precision': best_90pct['precision'],
                'recall': best_90pct['recall']
            })
        
        # Add analysis rows at the top
        if analysis_rows:
            analysis_df = pd.DataFrame(analysis_rows)
            threshold_df = pd.concat([analysis_df, threshold_df], ignore_index=True)
        
        return threshold_df
    
    def _log_summary_stats(self, ranked_df, threshold_df):
        """Log summary statistics"""
        
        total_flws = len(ranked_df)
        real_flws = (ranked_df['classification'] == 'real').sum()
        fake_flws = (ranked_df['classification'] == 'fake').sum()
        
        self.log(f"\nSummary: {total_flws} FLWs analyzed ({real_flws} real, {fake_flws} fake)")
        
        # Most suspicious real FLW
        most_suspicious_real = ranked_df[ranked_df['classification'] == 'real'].iloc[0]
        self.log(f"Most suspicious real FLW: {most_suspicious_real['flw_id']} (rank {most_suspicious_real['fraud_rank']}, score {most_suspicious_real['composite_prob']:.3f})")
        
        # Threshold analysis highlights
        analysis_rows = threshold_df[threshold_df['analysis'].notna()]
        for _, row in analysis_rows.iterrows():
            self.log(f"{row['analysis']}: threshold={row['threshold']:.3f} ? {row['flagged_real_pct']:.1f}% real flagged, {row['caught_fake_pct']:.1f}% fake caught")
    
    def _format_excel(self, writer, ranked_df):
        """Apply Excel formatting"""
        
        try:
            from openpyxl.styles import PatternFill
            
            workbook = writer.book
            worksheet = writer.sheets["Ranked_FLWs_Median_Fold"]
            
            # Light yellow fill for fake FLWs
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            
            # Find classification column
            classification_col = None
            for col_idx, col_name in enumerate(ranked_df.columns, 1):
                if col_name == 'classification':
                    classification_col = col_idx
                    break
            
            if classification_col is not None:
                for row_idx, (_, row_data) in enumerate(ranked_df.iterrows(), 2):
                    if row_data['classification'] == 'fake':
                        for col_idx in range(1, len(ranked_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = yellow_fill
        except Exception as e:
            self.log(f"Note: Could not apply Excel formatting - {str(e)}")
    
    @classmethod
    def create_for_automation(cls, df, output_dir, n_folds=10, include_rf=True,
                            include_lr=True, balance_method='class_weight', 
                            random_state=42, features_for_ml='all'):
        """
        Create validator instance for automated pipeline use
        
        Args:
            df: DataFrame with aggregated FLW features
            output_dir: Directory for output files
            n_folds: Number of folds (default 10)
            include_rf: Whether to train Random Forest
            include_lr: Whether to train Logistic Regression
            balance_method: Class imbalance handling
            random_state: Random seed
            features_for_ml: Either "all" or list of feature names to use
        
        Returns:
            MLKFoldValidator instance ready to call run_kfold_validation()
        """
        def dummy_log(msg):
            pass
        
        return cls(
            df=df,
            output_dir=output_dir,
            n_folds=n_folds,
            include_rf=include_rf,
            include_lr=include_lr,
            balance_method=balance_method,
            random_state=random_state,
            log_callback=dummy_log,
            features_for_ml=features_for_ml
        )
